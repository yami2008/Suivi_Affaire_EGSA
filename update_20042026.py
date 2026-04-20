"""
Mise à jour du 20/04/2026 — Planification semaine
"""
import openpyxl
from datetime import date

path = r'C:\Users\hp\Desktop\Suivi_Affaire_EGSA\Suivi_Affaires_EGSA.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

headers = [cell.value for cell in ws[1]]
col = {name: idx + 1 for idx, name in enumerate(headers)}

today_str = '20/04/2026'

updates = {
    'AFF-002': {
        'Date Prochaine Action': date(2026, 4, 25),
        'Date Alerte': date(2026, 4, 25),
        'Bloquant': 'Techniciens occupés — traitement prévu fin de semaine (25/04)',
        'historique': '20/04/2026 — Reporté à fin de semaine (25/04) : techniciens toujours occupés.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-003': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'RAPPEL demain après-midi 21/04 — Finaliser les lots CDC 2026 (AFF-004, 005, 006, 007 liées)',
        'historique': '20/04/2026 — Planifié pour demain après-midi 21/04/2026. Les lots AFF-004/005/006/007 seront traités en même temps.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-004': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'Traiter avec AFF-003 — demain après-midi 21/04/2026',
        'Tags': 'CDC2026, Lot1, PCs, LiéeAFF-003',
        'historique': '20/04/2026 — Sera traitée avec AFF-003 (CDC 2026) demain après-midi 21/04.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-005': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'Traiter avec AFF-003 — demain après-midi 21/04/2026',
        'Tags': 'CDC2026, Lot2, Architectes, LiéeAFF-003',
        'historique': '20/04/2026 — Sera traitée avec AFF-003 (CDC 2026) demain après-midi 21/04.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-006': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'Traiter avec AFF-003 — demain après-midi 21/04/2026',
        'Tags': 'CDC2026, Lot2, Laptops, Dev, LiéeAFF-003',
        'historique': '20/04/2026 — Sera traitée avec AFF-003 (CDC 2026) demain après-midi 21/04.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-007': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'Traiter avec AFF-003 — demain après-midi 21/04/2026',
        'Tags': 'CDC2026, Lot3, Lot4, SSD, LiéeAFF-003',
        'historique': '20/04/2026 — Sera traitée avec AFF-003 (CDC 2026) demain après-midi 21/04.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-009': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'RAPPEL demain matin 21/04 — Analyser BD ERP avec collègues + préparer résumé commercial réunion 14/04',
        'historique': '20/04/2026 — Planifié pour demain matin 21/04/2026 : analyse structure BD + résumé commercial réunion du 14/04.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-011': {
        'Statut': 'En cours',
        'Prochaine Action': "Rédiger résumé métier commercial (réunion 15/04) + intégrer à la conception ERP — en coordination avec AFF-009",
        'Bloquant': 'Nécessite coordination avec AFF-009 — conception ERP à faire ensemble',
        'Tags': 'Commercial, Réunion, ERP, ConceptionERP, LiéeAFF-009',
        'historique': "20/04/2026 — Réunion du 15/04 avec le commercial confirmée. Prochaine étape : rédiger résumé métier + l'intégrer à la conception ERP (lié AFF-009). Objectif : conception ERP complète et rigoureuse.",
        'Date Dernière MAJ': date(2026, 4, 20),
    },
    'AFF-013': {
        'Date Prochaine Action': date(2026, 4, 21),
        'Date Alerte': date(2026, 4, 21),
        'Prochaine Action': 'RAPPEL demain 21/04 — Relancer la directrice pour envoi des bilans mensuels J/F/M',
        'historique': '20/04/2026 — Rappel programmé pour demain 21/04/2026.',
        'Date Dernière MAJ': date(2026, 4, 20),
    },
}

for row in ws.iter_rows(min_row=2):
    aff_id = row[col['ID'] - 1].value
    if aff_id not in updates:
        continue

    u = updates[aff_id]
    for field, value in u.items():
        if field == 'historique':
            hist_cell = row[col['Historique'] - 1]
            existing = hist_cell.value or ''
            if value not in existing:
                hist_cell.value = (existing + '\n' + value).strip()
        elif field in col:
            row[col[field] - 1].value = value

    print(f'  OK {aff_id} mis a jour')

wb.save(path)
print()
print('Fichier sauvegarde avec succes.')
