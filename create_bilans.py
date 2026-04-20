#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from datetime import date, timedelta

# Charger le fichier Excel
wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

today = date(2026, 4, 20)

# Trouver le prochain numéro d'affaire
max_num = 0
for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    if not id_val:
        break
    if str(id_val).startswith('AFF-'):
        num = int(str(id_val).split('-')[1])
        max_num = max(max_num, num)

next_id = f'AFF-{max_num + 1:03d}'
next_row = ws.max_row + 1

print(f"Création de nouvelle affaire: {next_id}")
print()

# Ajouter la nouvelle affaire
ws[f'A{next_row}'] = next_id
ws[f'B{next_row}'] = 'Bilans mensuels — Janvier, Février, Mars'
ws[f'C{next_row}'] = 'Reporting'
ws[f'D{next_row}'] = 'La directrice doit envoyer les trois bilans mensuels (J/F/M) — État des lieux'
ws[f'E{next_row}'] = 'À traiter'
ws[f'F{next_row}'] = 'Haute'
ws[f'G{next_row}'] = 'Envoyer les trois bilans mensuels (janvier, février, mars) — État des lieux'
ws[f'H{next_row}'] = 'Directrice'
ws[f'I{next_row}'] = today  # Date ouverture
# J = Date limite (pas fixée)
ws[f'K{next_row}'] = today  # Date dernière MAJ
ws[f'L{next_row}'] = f'20/04/2026 — CRÉÉ : Affaire de suivi pour les trois bilans mensuels\nRappel quotidien jusqu\'à complération'
ws[f'Q{next_row}'] = today + timedelta(days=1)  # Première alerte demain
ws[f'R{next_row}'] = 'Imprévu'
ws[f'S{next_row}'] = 'Oui'  # À rappeler tous les jours
ws[f'T{next_row}'] = today  # Date prochaine action = aujourd'hui

# Sauvegarder
wb.save('Suivi_Affaires_EGSA.xlsx')

print(f"✅ {next_id} créé avec succès !")
print()
print("Détails :")
print(f"  Titre : Bilans mensuels — Janvier, Février, Mars")
print(f"  Responsable : Directrice")
print(f"  Statut : À traiter")
print(f"  Priorité : Haute")
print(f"  Imprévu : Oui (rappel quotidien jusqu'à complération)")
print(f"  Date prochaine action : {today.strftime('%d/%m/%Y')}")
