#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, date

# Charger le fichier Excel
wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

# Date du jour
today = date(2026, 4, 20)

print("=" * 80)
print("BRIEFING DU JOUR - " + today.strftime("%d/%m/%Y"))
print("=" * 80)
print()

# Dictionnaire pour stocker les affaires par type
urgences = []
alertes = []
prochaines_actions_auj = []
imprevu = []

# Itérer sur les lignes
for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    if not id_val:
        break
    
    titre = ws[f'B{row_idx}'].value
    type_aff = ws[f'C{row_idx}'].value
    desc = ws[f'D{row_idx}'].value
    statut = ws[f'E{row_idx}'].value
    priorite = ws[f'F{row_idx}'].value
    prox_action = ws[f'G{row_idx}'].value
    responsable = ws[f'H{row_idx}'].value
    date_ouv = ws[f'I{row_idx}'].value
    date_lim = ws[f'J{row_idx}'].value
    date_maj = ws[f'K{row_idx}'].value
    histo = ws[f'L{row_idx}'].value
    obs = ws[f'M{row_idx}'].value
    date_clot = ws[f'N{row_idx}'].value
    temps = ws[f'O{row_idx}'].value
    dossier = ws[f'P{row_idx}'].value
    date_alerte = ws[f'Q{row_idx}'].value
    origine = ws[f'R{row_idx}'].value
    imprevue = ws[f'S{row_idx}'].value
    date_prox_action = ws[f'T{row_idx}'].value
    bloquant = ws[f'U{row_idx}'].value
    tags = ws[f'V{row_idx}'].value
    semaine = ws[f'W{row_idx}'].value
    
    # Passer les affaires clôturées
    if statut == "Clôturée":
        continue
    
    affaire = {
        'id': id_val,
        'titre': titre,
        'statut': statut,
        'priorite': priorite,
        'prox_action': prox_action,
        'responsable': responsable,
        'date_alerte': date_alerte,
        'date_prox_action': date_prox_action,
        'imprevue': imprevue,
        'bloquant': bloquant,
        'obs': obs
    }
    
    # Classer par urgence
    # Convertir datetime en date si nécessaire
    date_alerte_cmp = date_alerte.date() if isinstance(date_alerte, datetime) else date_alerte
    date_prox_action_cmp = date_prox_action.date() if isinstance(date_prox_action, datetime) else date_prox_action
    
    # 1. Alertes dépassées (date alerte < aujourd'hui)
    if date_alerte_cmp and isinstance(date_alerte_cmp, date):
        if date_alerte_cmp < today:
            urgences.append(affaire)
            continue
        elif (date_alerte_cmp - today).days <= 3:
            alertes.append(affaire)
            continue
    
    # 2. Prochaine action = aujourd'hui
    if date_prox_action_cmp and isinstance(date_prox_action_cmp, date):
        if date_prox_action_cmp == today:
            prochaines_actions_auj.append(affaire)
            continue
    
    # 3. Tâches imprévues
    if imprevue == "Oui":
        imprevu.append(affaire)

# Affichage
print()
print("🚨 URGENCES À TRAITER IMMÉDIATEMENT (" + str(len(urgences)) + ")")
print("-" * 80)
for aff in urgences:
    print(f"  {aff['id']} | {aff['titre']}")
    print(f"    Statut: {aff['statut']} | Priorité: {aff['priorite']}")
    print(f"    Prochaine action: {aff['prox_action']}")
    print(f"    Responsable: {aff['responsable']}")
    if aff['bloquant']:
        print(f"    ⚠️  BLOQUANT: {aff['bloquant']}")
    print()

print()
print("⏰ ALERTES (< 3 jours) (" + str(len(alertes)) + ")")
print("-" * 80)
for aff in alertes:
    print(f"  {aff['id']} | {aff['titre']}")
    print(f"    Statut: {aff['statut']} | Priorité: {aff['priorite']}")
    print(f"    Prochaine action: {aff['prox_action']}")
    print(f"    Date alerte: {aff['date_alerte']}")
    print()

print()
print("📌 ACTIONS D'AUJOURD'HUI (" + str(len(prochaines_actions_auj)) + ")")
print("-" * 80)
for aff in prochaines_actions_auj:
    print(f"  {aff['id']} | {aff['titre']}")
    print(f"    Prochaine action: {aff['prox_action']}")
    print(f"    Responsable: {aff['responsable']}")
    if aff['obs']:
        print(f"    Notes: {aff['obs']}")
    print()

print()
print("🔧 TÂCHES IMPRÉVUES (" + str(len(imprevu)) + ")")
print("-" * 80)
for aff in imprevu:
    print(f"  {aff['id']} | {aff['titre']}")
    print(f"    Statut: {aff['statut']} | Priorité: {aff['priorite']}")
    if aff['obs']:
        print(f"    Notes: {aff['obs']}")
    print()

print()
print("=" * 80)
print(f"RÉSUMÉ: {len(urgences)} urgences + {len(alertes)} alertes + {len(prochaines_actions_auj)} actions = {len(urgences) + len(alertes) + len(prochaines_actions_auj)} tâches prioritaires")
print("=" * 80)
