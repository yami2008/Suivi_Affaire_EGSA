#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Confirmation plan 22/04/2026 :
AFF-011 & AFF-009 : Continuer résumé module Commercial + analyse base de données
"""

import openpyxl
from datetime import datetime

FILE = r'c:\Users\hp\Desktop\Suivi_Affaire_EGSA\Suivi_Affaires_EGSA.xlsx'
TODAY = datetime(2026, 4, 21, 0, 0)
DEMAIN = datetime(2026, 4, 22, 0, 0)
TODAY_STR = "21/04/2026"

wb = openpyxl.load_workbook(FILE)
ws = wb.active

rows = {}
for row_idx in range(2, ws.max_row + 1):
    val = ws[f'A{row_idx}'].value
    if val in ('AFF-009', 'AFF-011'):
        rows[val] = row_idx

print("Lignes trouvées :", rows)

PROX_ACTION = "À CONTINUER demain 22/04 — Résumé module Commercial + analyse structure base de données (en cours depuis 21/04)"

for aff_id in ('AFF-011', 'AFF-009'):
    r = rows[aff_id]
    ws[f'G{r}'].value = PROX_ACTION
    ws[f'K{r}'].value = TODAY
    ws[f'Q{r}'].value = DEMAIN
    ws[f'T{r}'].value = DEMAIN

    hist = ws[f'L{r}'].value or ''
    hist += f"\n{TODAY_STR} — Plan confirmé pour 22/04 : continuer résumé module Commercial + analyse structure BD. Travail en cours."
    ws[f'L{r}'].value = hist

wb.save(FILE)
print("✅ Mis à jour !")
print(f"  AFF-011 & AFF-009 → Prochaine action : {PROX_ACTION}")
