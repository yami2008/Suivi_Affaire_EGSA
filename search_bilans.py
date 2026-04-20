#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

print('Recherche des affaires concernant les bilans (janvier, février, mars)...\n')

for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    titre = ws[f'B{row_idx}'].value
    statut = ws[f'E{row_idx}'].value
    priorite = ws[f'F{row_idx}'].value
    prox_action = ws[f'G{row_idx}'].value
    responsable = ws[f'H{row_idx}'].value
    date_lim = ws[f'J{row_idx}'].value
    date_alerte = ws[f'Q{row_idx}'].value
    date_maj = ws[f'K{row_idx}'].value
    obs = ws[f'M{row_idx}'].value
    
    if not id_val:
        break
    
    keywords = ['bilan', 'état des lieux', 'janvier', 'février', 'mars', 'J/F/M', 'trimestre', 'mensuel']
    if titre and any(kw in titre.lower() for kw in keywords):
        print(f'{id_val} | {titre}')
        print(f'  Statut: {statut} | Priorité: {priorite}')
        print(f'  Responsable: {responsable}')
        print(f'  Date limite: {date_lim}')
        print(f'  Date alerte: {date_alerte}')
        print(f'  Dernière MAJ: {date_maj}')
        print(f'  Prochaine action: {prox_action}')
        if obs:
            print(f'  Observations: {obs}')
        print()
