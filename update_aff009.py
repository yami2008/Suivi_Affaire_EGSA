#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from datetime import date

wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

today = date(2026, 4, 20)

for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    if not id_val:
        break

    if id_val == 'AFF-009':
        # Prochaine action : deux tâches
        ws[f'G{row_idx}'] = (
            '1) Analyser la structure de la BD ERP (tables, champs, relations) '
            '— faire avec les collègues\n'
            '2) Préparer le résumé commercial de la réunion de mercredi 15/04'
        )
        # Lever le bloquant (BD identifiée et restaurée)
        ws[f'U{row_idx}'] = None
        # Historique
        histo = ws[f'L{row_idx}'].value or ''
        histo += (
            '\n20/04/2026 — Base de données restaurée en local (.bak OK). '
            'BD identifiée : base des commerciaux en production. '
            'Prochaine étape : analyse structure + résumé commercial (en groupe avec collègues).'
        )
        ws[f'L{row_idx}'] = histo
        # Statut
        ws[f'E{row_idx}'] = 'En cours'
        # Date dernière MAJ
        ws[f'K{row_idx}'] = today
        print('AFF-009 mis à jour :')
        print('  ✓ Bloquant levé (BD restaurée)')
        print('  ✓ Prochaine action : Analyser BD + Résumé commercial')
        print('  ✓ Historique mis à jour')
        break

wb.save('Suivi_Affaires_EGSA.xlsx')
print('\n✅ Fichier sauvegardé')
