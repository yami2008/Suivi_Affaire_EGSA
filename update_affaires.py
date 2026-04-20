#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, date

# Charger le fichier Excel
wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

today = date(2026, 4, 20)
print(f"Mise à jour des affaires - {today.strftime('%d/%m/%Y')}\n")

# Trouver les lignes des affaires à mettre à jour
for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    if not id_val:
        break
    
    if id_val == 'AFF-012':
        print(f"AFF-012 : Clôture des décharges")
        # Colonne E = Statut
        ws[f'E{row_idx}'] = 'Clôturée'
        # Colonne N = Date Clôture
        ws[f'N{row_idx}'] = today
        # Colonne G = Prochaine action
        ws[f'G{row_idx}'] = 'Affaire clôturée — Décharges signées par 3 comptables, 8 mini PC envoyés (3 aéroports)'
        # Colonne L = Historique
        histo = ws[f'L{row_idx}'].value or ""
        histo += f"\n20/04/2026 — CLÔTURÉ : Décharges signées, 8 mini PC envoyés dans 3 aéroports"
        ws[f'L{row_idx}'] = histo
        # Colonne K = Date Dernière MAJ
        ws[f'K{row_idx}'] = today
        print("  ✓ Statut = Clôturée")
        print("  ✓ Historique mis à jour\n")
    
    elif id_val == 'AFF-002':
        print(f"AFF-002 : Réparation Netbox Aéroport")
        # Déplacer la date alerte au jeudi prochain (23/04/2026)
        next_thursday = date(2026, 4, 23)
        ws[f'Q{row_idx}'] = next_thursday
        # Colonne L = Historique
        histo = ws[f'L{row_idx}'].value or ""
        histo += f"\n20/04/2026 — Délai prolongé au jeudi 23/04/2026 (techniciens occupés)"
        ws[f'L{row_idx}'] = histo
        # Colonne K = Date Dernière MAJ
        ws[f'K{row_idx}'] = today
        print("  ✓ Date alerte = 23/04/2026")
        print("  ✓ Historique mis à jour\n")
    
    elif id_val == 'AFF-009':
        print(f"AFF-009 : Analyse base de données ERP")
        # Note : pas encore commencé, dimanche sans bureau
        # Colonne L = Historique
        histo = ws[f'L{row_idx}'].value or ""
        histo += f"\n19/04/2026 — Jour sans bureau (dimanche), aucune action possibility"
        histo += f"\n20/04/2026 — A COMMENCER : appeler collègue pour identifier BD production"
        ws[f'L{row_idx}'] = histo
        # Colonne K = Date Dernière MAJ
        ws[f'K{row_idx}'] = today
        print("  ✓ Historique mis à jour (pas commencé, à faire aujourd'hui)")
        print("  ✓ Statut reste En cours\n")
    
    elif id_val == 'AFF-011':
        print(f"AFF-011 : Entretien avec le commercial")
        # La réunion a été faite
        # Colonne G = Prochaine action
        ws[f'G{row_idx}'] = 'Réunion effectuée — à clôturer ou reporter selon outcome'
        # Colonne L = Historique
        histo = ws[f'L{row_idx}'].value or ""
        histo += f"\n20/04/2026 — Réunion avec commercial effectuée"
        ws[f'L{row_idx}'] = histo
        # Colonne K = Date Dernière MAJ
        ws[f'K{row_idx}'] = today
        # Colonne E = Statut → En cours (réunion faite, en attente de suites)
        ws[f'E{row_idx}'] = 'En cours'
        print("  ✓ Réunion effectuée")
        print("  ✓ Statut = En cours")
        print("  ✓ Historique mis à jour\n")

# Sauvegarder
wb.save('Suivi_Affaires_EGSA.xlsx')
print("✅ Fichier sauvegardé avec succès")
