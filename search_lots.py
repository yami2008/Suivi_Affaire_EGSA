#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

print('Recherche des affaires contenant "lots"...\n')

for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    titre = ws[f'B{row_idx}'].value
    statut = ws[f'E{row_idx}'].value
    date_lim = ws[f'J{row_idx}'].value
    prox_action = ws[f'G{row_idx}'].value
    desc = ws[f'D{row_idx}'].value
    
    if not id_val:
        break
    
    if titre and 'lot' in titre.lower():
        print(f'{id_val} | {titre}')
        print(f'  Statut: {statut}')
        print(f'  Description: {desc}')
        print(f'  Date limite: {date_lim}')
        print(f'  Prochaine action: {prox_action}')
        print()
