#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Mises à jour du 21/04/2026 :
- AFF-011 : Démarrage ce matin (peut se prolonger jusqu'à demain)
- AFF-009 : Traitement en parallèle avec AFF-011 ce matin
- AFF-003/004/005/006 : Traitement cet après-midi
- AFF-013 : Rappel cet après-midi
"""

import openpyxl
from datetime import datetime

FILE = r'c:\Users\hp\Desktop\Suivi_Affaire_EGSA\Suivi_Affaires_EGSA.xlsx'
TODAY = datetime(2026, 4, 21, 0, 0)
TODAY_STR = "21/04/2026"

wb = openpyxl.load_workbook(FILE)
ws = wb.active

# Trouver les lignes de chaque affaire
rows = {}
for row_idx in range(2, ws.max_row + 1):
    val = ws[f'A{row_idx}'].value
    if val in ('AFF-003', 'AFF-004', 'AFF-005', 'AFF-006', 'AFF-009', 'AFF-011', 'AFF-013'):
        rows[val] = row_idx

print("Lignes trouvées :", rows)

# ─────────────────────────────────────────────
# AFF-011 : Démarrage ce matin
# ─────────────────────────────────────────────
r = rows['AFF-011']
ws[f'E{r}'].value = 'En cours'
ws[f'G{r}'].value = "EN COURS ce matin 21/04 — Rédiger résumé métier commercial (réunion 15/04) + intégrer à la conception ERP (avec AFF-009) — peut se prolonger jusqu'au 22/04"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_011 = ws[f'L{r}'].value or ''
hist_011 += f"\n{TODAY_STR} — DÉMARRÉ ce matin. Rédaction du résumé métier commercial + intégration conception ERP. En parallèle avec AFF-009. Peut se prolonger jusqu'au 22/04."
ws[f'L{r}'].value = hist_011

# ─────────────────────────────────────────────
# AFF-009 : En parallèle avec AFF-011 ce matin
# ─────────────────────────────────────────────
r = rows['AFF-009']
ws[f'E{r}'].value = 'En cours'
ws[f'G{r}'].value = "EN COURS ce matin 21/04 — Analyse BD ERP + résumé commercial réunion 14/04 — en parallèle avec AFF-011"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_009 = ws[f'L{r}'].value or ''
hist_009 += f"\n{TODAY_STR} — DÉMARRÉ ce matin, en parallèle avec AFF-011. Analyse structure BD + rédaction résumé commercial réunion 14/04."
ws[f'L{r}'].value = hist_009

# ─────────────────────────────────────────────
# AFF-003 : Cet après-midi
# ─────────────────────────────────────────────
r = rows['AFF-003']
ws[f'G{r}'].value = "AUJOURD'HUI après-midi 21/04 — Finaliser les lots CDC 2026 (AFF-004, 005, 006, 007 liées)"
ws[f'K{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_003 = ws[f'L{r}'].value or ''
hist_003 += f"\n{TODAY_STR} — Planifié pour cet après-midi 21/04. Finalisation des lots CDC 2026 avec AFF-004/005/006/007."
ws[f'L{r}'].value = hist_003

# ─────────────────────────────────────────────
# AFF-004 : Cet après-midi
# ─────────────────────────────────────────────
r = rows['AFF-004']
ws[f'G{r}'].value = "AUJOURD'HUI après-midi 21/04 — Traiter avec AFF-003 (Lot N°1 : PCs Bureau)"
ws[f'K{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_004 = ws[f'L{r}'].value or ''
hist_004 += f"\n{TODAY_STR} — Planifié pour cet après-midi 21/04 avec AFF-003."
ws[f'L{r}'].value = hist_004

# ─────────────────────────────────────────────
# AFF-005 : Cet après-midi
# ─────────────────────────────────────────────
r = rows['AFF-005']
ws[f'G{r}'].value = "AUJOURD'HUI après-midi 21/04 — Traiter avec AFF-003 (Lot N°2 : PCs Architectes)"
ws[f'K{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_005 = ws[f'L{r}'].value or ''
hist_005 += f"\n{TODAY_STR} — Planifié pour cet après-midi 21/04 avec AFF-003."
ws[f'L{r}'].value = hist_005

# ─────────────────────────────────────────────
# AFF-006 : Cet après-midi
# ─────────────────────────────────────────────
r = rows['AFF-006']
ws[f'G{r}'].value = "AUJOURD'HUI après-midi 21/04 — Traiter avec AFF-003 (Lot N°2 : Laptops Développeurs)"
ws[f'K{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_006 = ws[f'L{r}'].value or ''
hist_006 += f"\n{TODAY_STR} — Planifié pour cet après-midi 21/04 avec AFF-003."
ws[f'L{r}'].value = hist_006

# ─────────────────────────────────────────────
# AFF-013 : Rappel cet après-midi
# ─────────────────────────────────────────────
r = rows['AFF-013']
ws[f'G{r}'].value = "RAPPEL CET APRÈS-MIDI 21/04 — Relancer la directrice pour envoi des bilans mensuels J/F/M"
ws[f'K{r}'].value = TODAY
ws[f'T{r}'].value = TODAY

hist_013 = ws[f'L{r}'].value or ''
hist_013 += f"\n{TODAY_STR} — Rappel programmé pour cet après-midi 21/04/2026."
ws[f'L{r}'].value = hist_013

# ─────────────────────────────────────────────
# Sauvegarde
# ─────────────────────────────────────────────
wb.save(FILE)
print("✅ Fichier mis à jour avec succès !")
print("\nRécapitulatif des modifications :")
print("  AFF-011 : Démarré ce matin — peut se prolonger jusqu'au 22/04")
print("  AFF-009 : Démarré ce matin en parallèle avec AFF-011")
print("  AFF-003 : Cet après-midi")
print("  AFF-004 : Cet après-midi (avec AFF-003)")
print("  AFF-005 : Cet après-midi (avec AFF-003)")
print("  AFF-006 : Cet après-midi (avec AFF-003)")
print("  AFF-013 : Rappel cet après-midi")
