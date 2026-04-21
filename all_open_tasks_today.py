#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from datetime import date, datetime

wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

today = date(2026, 4, 21)

print("=" * 70)
print("TOUTES LES TÂCHES NON TERMINÉES — 21/04/2026")
print("=" * 70)

urgences = []
haute = []
moyenne = []
basse = []

for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    if not id_val:
        break

    statut = ws[f'E{row_idx}'].value
    if statut and 'Clôturée' in str(statut):
        continue

    titre = ws[f'B{row_idx}'].value
    priorite = ws[f'F{row_idx}'].value
    prox_action = ws[f'G{row_idx}'].value
    responsable = ws[f'H{row_idx}'].value
    date_lim = ws[f'J{row_idx}'].value
    date_alerte_raw = ws[f'Q{row_idx}'].value
    date_prox_raw = ws[f'T{row_idx}'].value
    bloquant = ws[f'U{row_idx}'].value
    obs = ws[f'M{row_idx}'].value

    # Normaliser les dates
    def to_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        return None

    date_alerte = to_date(date_alerte_raw)
    date_prox = to_date(date_prox_raw)

    aff = {
        'id': id_val,
        'titre': titre,
        'statut': statut,
        'priorite': priorite,
        'prox_action': prox_action,
        'responsable': responsable,
        'date_lim': date_lim,
        'date_alerte': date_alerte,
        'bloquant': bloquant,
        'obs': obs,
    }

    # Classer
    if date_alerte and date_alerte < today:
        urgences.append(aff)
    elif str(priorite).lower() == 'haute':
        haute.append(aff)
    elif str(priorite).lower() == 'moyenne':
        moyenne.append(aff)
    else:
        basse.append(aff)

def afficher(liste, titre_section):
    print(f"\n{titre_section} ({len(liste)})")
    print("-" * 70)
    if not liste:
        print("  Aucune")
        return
    for a in liste:
        print(f"  {a['id']} | {a['titre']}")
        print(f"    Statut: {a['statut']} | Priorité: {a['priorite']}")
        if a['prox_action']:
            print(f"    → {a['prox_action']}")
        if a['responsable']:
            print(f"    Responsable: {a['responsable']}")
        if a['bloquant']:
            print(f"    ⚠️  BLOQUANT: {a['bloquant']}")
        if a['date_alerte']:
            delta = (a['date_alerte'] - today).days
            if delta >= 0:
                print(f"    📅 Alerte dans {delta} jour(s) ({a['date_alerte'].strftime('%d/%m/%Y')})")
            else:
                print(f"    🔴 Alerte DÉPASSÉE de {-delta} jour(s) ({a['date_alerte'].strftime('%d/%m/%Y')})")
        print()

afficher(urgences, "🚨 URGENCES (alertes dépassées)")
afficher(haute, "🔴 PRIORITÉ HAUTE")
afficher(moyenne, "🟠 PRIORITÉ MOYENNE")
afficher(basse, "🟡 PRIORITÉ BASSE / NON DÉFINIE")

total = len(urgences) + len(haute) + len(moyenne) + len(basse)
print("=" * 70)
print(f"TOTAL : {total} tâche(s) non terminée(s)")
print("=" * 70)
