#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Mise à jour fin de journée 21/04/2026 :
- AFF-011 & AFF-009 : Démarrées, en cours (analyse BD + résumé commercial), pas terminées -> à continuer demain
- AFF-003/004/005/006/013 : Non traitées aujourd'hui -> report au 22/04
"""

import openpyxl
from datetime import datetime

FILE = r'c:\Users\hp\Desktop\Suivi_Affaire_EGSA\Suivi_Affaires_EGSA.xlsx'
TODAY = datetime(2026, 4, 21, 0, 0)
DEMAIN = datetime(2026, 4, 22, 0, 0)
TODAY_STR = "21/04/2026"
DEMAIN_STR = "22/04/2026"

wb = openpyxl.load_workbook(FILE)
ws = wb.active

# Trouver les lignes de chaque affaire
rows = {}
targets = ('AFF-003', 'AFF-004', 'AFF-005', 'AFF-006', 'AFF-009', 'AFF-011', 'AFF-013')
for row_idx in range(2, ws.max_row + 1):
    val = ws[f'A{row_idx}'].value
    if val in targets:
        rows[val] = row_idx

print("Lignes trouvées :", rows)

# ─────────────────────────────────────────────
# AFF-011 : En cours — pas terminée, continue demain
# ─────────────────────────────────────────────
r = rows['AFF-011']
ws[f'E{r}'].value = 'En cours'
ws[f'G{r}'].value = "À CONTINUER demain 22/04 — Analyse BD ERP + résumé module Commercial (commencé 21/04, pas terminé)"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Travail commencé : analyse BD + résumé module Commercial. Non terminé en fin de journée. À continuer le {DEMAIN_STR}."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# AFF-009 : En cours — pas terminée, continue demain
# ─────────────────────────────────────────────
r = rows['AFF-009']
ws[f'E{r}'].value = 'En cours'
ws[f'G{r}'].value = "À CONTINUER demain 22/04 — Analyse BD ERP + résumé module Commercial (commencé 21/04, pas terminé)"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Travail commencé : analyse BD + résumé module Commercial. Non terminé en fin de journée. À continuer le {DEMAIN_STR}."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# AFF-003 : Non traité aujourd'hui -> report 22/04
# ─────────────────────────────────────────────
r = rows['AFF-003']
ws[f'G{r}'].value = "REPORT au 22/04 — Finaliser les lots CDC 2026 (AFF-004, 005, 006, 007 liées) — non traité le 21/04"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Non traité (priorité donnée à AFF-011/009). Reporté au {DEMAIN_STR}."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# AFF-004 : Non traité aujourd'hui -> report 22/04
# ─────────────────────────────────────────────
r = rows['AFF-004']
ws[f'G{r}'].value = "REPORT au 22/04 — Traiter avec AFF-003 (Lot N°1 : PCs Bureau) — non traité le 21/04"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Non traité. Reporté au {DEMAIN_STR} avec AFF-003."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# AFF-005 : Non traité aujourd'hui -> report 22/04
# ─────────────────────────────────────────────
r = rows['AFF-005']
ws[f'G{r}'].value = "REPORT au 22/04 — Traiter avec AFF-003 (Lot N°2 : PCs Architectes) — non traité le 21/04"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Non traité. Reporté au {DEMAIN_STR} avec AFF-003."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# AFF-006 : Non traité aujourd'hui -> report 22/04
# ─────────────────────────────────────────────
r = rows['AFF-006']
ws[f'G{r}'].value = "REPORT au 22/04 — Traiter avec AFF-003 (Lot N°2 : Laptops Développeurs) — non traité le 21/04"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Non traité. Reporté au {DEMAIN_STR} avec AFF-003."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# AFF-013 : Non traité aujourd'hui -> rappel demain
# ─────────────────────────────────────────────
r = rows['AFF-013']
ws[f'G{r}'].value = "RAPPEL demain 22/04 — Relancer la directrice pour envoi des bilans mensuels J/F/M — non traité le 21/04"
ws[f'K{r}'].value = TODAY
ws[f'Q{r}'].value = DEMAIN
ws[f'T{r}'].value = DEMAIN

hist = ws[f'L{r}'].value or ''
hist += f"\n{TODAY_STR} — Non traité. Rappel reporté au {DEMAIN_STR}."
ws[f'L{r}'].value = hist

# ─────────────────────────────────────────────
# Sauvegarde
# ─────────────────────────────────────────────
wb.save(FILE)
print("✅ Fichier mis à jour avec succès !")
print("\nRécapitulatif :")
print("  AFF-011 : En cours — à continuer le 22/04")
print("  AFF-009 : En cours — à continuer le 22/04")
print("  AFF-003 : Reporté au 22/04")
print("  AFF-004 : Reporté au 22/04")
print("  AFF-005 : Reporté au 22/04")
print("  AFF-006 : Reporté au 22/04")
print("  AFF-013 : Rappel reporté au 22/04")
