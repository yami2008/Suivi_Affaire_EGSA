import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date

path = r'C:\Users\hp\Desktop\YYYYYYY\Suivi_Affaire_EGSA\Suivi_Affaires_EGSA.xlsx'
wb   = openpyxl.load_workbook(path)
ws   = wb['Suivi Affaires']

# ── Styles ────────────────────────────────────────────────────────────────
thin         = Side(border_style="thin", color="CCCCCC")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT_HEADER  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FONT_CELL    = Font(name="Arial", size=10)
FILL_HEADER  = PatternFill("solid", fgColor="2E4057")
ROUGE        = PatternFill("solid", fgColor="FF4444")
ORANGE       = PatternFill("solid", fgColor="FFA500")
GRIS         = PatternFill("solid", fgColor="D9D9D9")

# ── Dossiers liés aux affaires ─────────────────────────────────────────────
# Ajouter une entrée ici quand un nouveau dossier est créé
dossiers = {
    'AFF-001': r'C:\Users\hp\Desktop\YYYYYYY\Suivi_Affaire_EGSA\AFF-001_BIG_Informatique',
}

# ── Mises à jour du jour ───────────────────────────────────────────────────
# Format : 'AFF-XXX': {'historique_ajout': '...', 'maj': '...', 'temps': '...'}
mises_a_jour = {
    'AFF-008': {
        'historique_ajout': '14/04 : Assistance à la réunion commerciale dans la salle de formation',
        'maj': '2026-04-14',
        'temps': 'Matinée 12/04 (~4h) + installation 13/04 + assistance réunion 14/04',
    },
    'AFF-009': {
        'historique_ajout': '14/04 : Base de données .bak reçue de la collègue — analyse à démarrer en urgence',
        'maj': '2026-04-14',
    },
}

# ── Récupérer les en-têtes ─────────────────────────────────────────────────
headers = [cell.value for cell in ws[1]]
col      = {name: idx + 1 for idx, name in enumerate(headers)}

# ── Appliquer hyperliens + styles par ligne ────────────────────────────────
today = date.today()

for row in ws.iter_rows(min_row=2):
    aff_id     = row[col['ID'] - 1].value
    statut     = row[col['Statut'] - 1].value or ''
    alerte_val = row[col['Date Alerte'] - 1].value if 'Date Alerte' in col else None

    # Mise à jour historique/données
    if aff_id in mises_a_jour:
        update = mises_a_jour[aff_id]
        if 'historique_ajout' in update:
            hist_cell = row[col['Historique'] - 1]
            existing = hist_cell.value or ''
            entry = update['historique_ajout']
            if entry not in existing:
                hist_cell.value = existing + '\n' + entry if existing else entry
        if 'maj' in update:
            row[col['Date Dernière MAJ'] - 1].value = update['maj']
        if 'temps' in update:
            row[col['Temps passé'] - 1].value = update['temps']

    # Hyperlien dossier
    if aff_id in dossiers:
        cell_dossier = row[col['Dossier'] - 1]
        dossier_path = dossiers[aff_id]
        cell_dossier.value     = dossier_path
        cell_dossier.hyperlink = 'file:///' + dossier_path.replace('\\', '/')
        cell_dossier.font      = Font(name="Arial", size=10, color="0563C1", underline="single")

    # Colorisation ligne selon alerte/statut
    cloturee   = 'clôtur' in statut.lower()
    fill_ligne = None
    if cloturee:
        fill_ligne = GRIS
    elif alerte_val:
        try:
            alerte_date = datetime.strptime(str(alerte_val), '%Y-%m-%d').date()
            if alerte_date <= today:
                fill_ligne = ROUGE
            elif (alerte_date - today).days <= 3:
                fill_ligne = ORANGE
        except Exception:
            pass

    # Appliquer style à chaque cellule de la ligne
    for cell in row:
        if cell.column != col.get('Dossier'):  # garder le style hyperlien sur Dossier
            cell.font = FONT_CELL
        cell.border    = BORDER
        cell.alignment = Alignment(wrap_text=False, vertical='center')
        if fill_ligne:
            cell.fill = fill_ligne

# ── En-tête ────────────────────────────────────────────────────────────────
for cell in ws[1]:
    cell.font      = FONT_HEADER
    cell.fill      = FILL_HEADER
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = BORDER

ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

# ── Largeurs colonnes ──────────────────────────────────────────────────────
largeurs = {
    'ID': 10, 'Titre': 36, 'Type': 20, 'Description': 42,
    'Statut': 28, 'Priorité': 12, 'Prochaine Action': 42,
    'Responsable': 14, 'Date Ouverture': 16, 'Date Limite': 14,
    'Date Dernière MAJ': 18, 'Historique': 55, 'Observations': 42,
    'Date Clôture': 14, 'Temps passé': 22, 'Dossier': 55,
    'Date Alerte': 14, 'Origine': 24, 'Imprévu': 10,
    'Date Prochaine Action': 22, 'Bloquant': 45, 'Tags': 30, 'Semaine N°': 13,
}
for i, h in enumerate(headers, start=1):
    if h in largeurs:
        ws.column_dimensions[get_column_letter(i)].width = largeurs[h]

wb.save(path)
print('Fichier sauvegardé — hyperliens + styles appliqués.')
