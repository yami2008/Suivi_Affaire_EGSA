#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('Suivi_Affaires_EGSA.xlsx')
ws = wb.active

print('=== AFFAIRES CDC 2026 (LOTS) ===\n')

cdc_ids = ['AFF-004', 'AFF-005', 'AFF-006', 'AFF-007']

for row_idx in range(2, ws.max_row + 1):
    id_val = ws[f'A{row_idx}'].value
    if not id_val:
        break
    
    if id_val in cdc_ids:
        titre = ws[f'B{row_idx}'].value
        statut = ws[f'E{row_idx}'].value
        priorite = ws[f'F{row_idx}'].value
        prox_action = ws[f'G{row_idx}'].value
        responsable = ws[f'H{row_idx}'].value
        desc = ws[f'D{row_idx}'].value
        date_maj = ws[f'K{row_idx}'].value
        obs = ws[f'M{row_idx}'].value
        
        print(f'{id_val} | {titre}')
        print(f'  Statut: {statut} | Priorité: {priorite}')
        print(f'  Responsable: {responsable}')
        print(f'  Description: {desc}')
        print(f'  Prochaine action: {prox_action}')
        print(f'  Dernière MAJ: {date_maj}')
        if obs:
            print(f'  Observations: {obs}')
        print()
